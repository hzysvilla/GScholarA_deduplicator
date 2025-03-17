import os
import datetime
import base64
import logging
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from bs4 import BeautifulSoup
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
import google.auth.exceptions
from collections import defaultdict
from termcolor import colored

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Constants
SCOPES = ['https://www.googleapis.com/auth/gmail.modify']
EMAIL_QUERY = 'from:scholaralerts-noreply@google.com is:unread'
HISTORY_FOLDER = os.path.join(os.getcwd(), 'history')
MAX_DUPNUM=256


COLOR_PALETTE = [
    PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),  # Light gray
    PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid"),  # Light blue
    PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"),  # Light pink
    PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid"),  # Light green
]

def get_email_body(payload):
    """Extract the HTML body from the email payload."""
    body = ""
    if 'parts' in payload:
        for part in payload['parts']:
            if part['mimeType'] == 'text/html':
                body = base64.urlsafe_b64decode(part['body']['data']).decode('utf-8')
                break
    else:
        if payload['mimeType'] == 'text/html':
            body = base64.urlsafe_b64decode(payload['body']['data']).decode('utf-8')
    return body

def parse_email_html(html_content):
    """Parse the HTML content to extract paper details."""
    soup = BeautifulSoup(html_content, 'html.parser')
    papers = []
    
    for item in soup.find_all('div', style=lambda x: x and 'line-height:17px' in x):
        title_tag = item.find_previous('a', class_='gse_alrt_title')
        authors_tag = item.find_previous('div', style=lambda x: x and 'color:#006621' in x)
        snippet = item.text.strip()
        
        if title_tag and authors_tag:
            title = title_tag.text.strip()
            authors = authors_tag.text.strip()
            papers.append({
                'title': title,
                'authors': authors,
                'snippet': snippet
            })
    
    return papers

def wrap_snippet(snippet, words_per_line=13):
    """Wrap the snippet text after every specified number of words."""
    words = snippet.split()
    wrapped_lines = []
    for i in range(0, len(words), words_per_line):
        wrapped_lines.append(' '.join(words[i:i + words_per_line]))
    return '\n'.join(wrapped_lines)

def save_to_excel(papers, email_count):
    """Save the extracted papers to an Excel file with improved formatting."""
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H-%M-%S")
    os.makedirs(HISTORY_FOLDER, exist_ok=True)
    file_name = f"{timestamp}_{email_count}_emails.xlsx"
    file_path = os.path.join(HISTORY_FOLDER, file_name)
    
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scholar Alerts"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin'))
    alignment = Alignment(wrap_text=True, vertical='top')
    
    # Write headers
    headers = [f"Paper Details (Total: {len(papers)})"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = alignment
    
    # Write paper details
    row_offset = 2  # Start from row 2 (after header)
    max_snippet_width = 0  # Track the maximum width needed for snippets
    
    for i, paper in enumerate(papers):
        fill_color = COLOR_PALETTE[i % len(COLOR_PALETTE)]
        
        # Title
        title_cell = ws.cell(row=row_offset, column=1, value=f"Title{i+1}: {paper['title']} ({paper['count']})")
        title_cell.border = border
        title_cell.alignment = alignment
        
        # Authors
        authors_cell = ws.cell(row=row_offset + 1, column=1, value=f"Authors{i+1}: {paper['authors']}")
        authors_cell.border = border
        authors_cell.alignment = alignment
        
        # Snippet
        wrapped_snippet = wrap_snippet(paper['snippet'])
        snippet_cell = ws.cell(row=row_offset + 2, column=1, value=f"Snippet{i+1}: {wrapped_snippet}")
        snippet_cell.border = border
        snippet_cell.alignment = alignment
        
        # Calculate the number of lines in the snippet
        lines = str(snippet_cell.value).count('\n') + 1
        ws.row_dimensions[row_offset + 2].height = lines * 15  # Adjust height based on number of lines
        
        # Update the maximum width needed for snippets
        snippet_lines = str(snippet_cell.value).split('\n')
        for line in snippet_lines:
            max_snippet_width = max(max_snippet_width, len(line))
        
        # Apply background color to each cell
        for row in range(row_offset, row_offset + 3):
            ws.cell(row=row, column=1).fill = fill_color
        
        row_offset += 4  # Add extra row for spacing
    
    # Auto-adjust column width based on the longest snippet line
    adjusted_width = max_snippet_width + 2
    ws.column_dimensions[get_column_letter(1)].width = adjusted_width
    
    # Ensure the column width is at least 50 characters wide
    if adjusted_width < 50:
        ws.column_dimensions[get_column_letter(1)].width = 50
    
    # Save the workbook
    wb.save(file_path)
    logging.info(f"Saved results to {file_path}")

def main():
    """Main function to fetch emails, parse content, and save to Excel."""
    personal_info = "Hello Zheyuan He (ecjgvmhc@gmail.com, KTH)"
    print(colored(personal_info, 'green'))

    creds = None
    token_path = 'token.json'

    try:
        if os.path.exists(token_path):
            creds = Credentials.from_authorized_user_file(token_path, SCOPES)

        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                try:
                    creds.refresh(Request())
                except google.auth.exceptions.RefreshError:
                    logging.error("Token has expired or been revoked. Deleting token.json and retrying...")
                    os.remove(token_path)
                    return main()  # Restart the program

            else:
                flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
                creds = flow.run_local_server(port=0)

            with open(token_path, 'w') as token:
                token.write(creds.to_json())

    except Exception as e:
        logging.error(f"Unexpected error during authentication: {e}")
        return

    service = build('gmail', 'v1', credentials=creds)
    results = service.users().messages().list(userId='me', q=EMAIL_QUERY, maxResults=MAX_DUPNUM).execute()
    messages = results.get('messages', [])

    if not messages:
        logging.info('No unread messages found from scholaralerts-noreply@google.com.')
        return

    all_papers = []
    for message in messages:
        msg = service.users().messages().get(userId='me', id=message['id'], format='full').execute()
        html_content = get_email_body(msg['payload'])
        if html_content:
            papers = parse_email_html(html_content)
            all_papers.extend(papers)

        service.users().messages().modify(
            userId='me',
            id=message['id'],
            body={'removeLabelIds': ['UNREAD']}
        ).execute()
        logging.info(f"Marked message {message['id']} as read.")

    # Count occurrences of each paper
    paper_counts = defaultdict(int)
    for paper in all_papers:
        key = (paper['title'], paper['authors'])
        paper_counts[key] += 1

    # Create a list of unique papers with their counts
    unique_papers = []
    for (title, authors), count in paper_counts.items():
        for paper in all_papers:
            if paper['title'] == title and paper['authors'] == authors:
                unique_papers.append({'title': title, 'authors': authors, 'snippet': paper['snippet'], 'count': count})
                break

    # Sort papers by count in descending order
    unique_papers.sort(key=lambda x: x['count'], reverse=True)

    save_to_excel(unique_papers, len(messages))


if __name__ == '__main__':
    main()

